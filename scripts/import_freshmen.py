"""
大一学生数据导入脚本
读取大一名单.xlsx和大一部分班级班级干部任命(2).xlsx
创建用户账号和报名记录
"""

import openpyxl
import sys
import os
import asyncio
import bcrypt

sys.stdout.reconfigure(encoding='utf-8')

# 设置环境变量
os.environ['DATABASE_URL'] = 'postgresql://postgres:FidKxGIyuBSQ35w6@db.eqergssmphvluywyysif.supabase.co:5432/postgres'

from prisma import Prisma

POSITION_KEYWORDS = [
    ('班长', 'CLASS_MONITOR', 1),
    ('团支书', 'LEAGUE_SECRETARY', 2),
    ('学习委员', 'STUDY_COMMISSAR', 3),
    ('学委', 'STUDY_COMMISSAR', 3),
    ('生活班长', 'LIFE_COMMISSAR', 4),
    ('生活委员', 'LIFE_COMMISSAR', 4),
    ('文体委员', 'CULTURE_COMMISSAR', 5),
    ('文艺委员', 'CULTURE_COMMISSAR', 5),
    ('心理委员', 'NONE', 6),
    ('宣传委员', 'NONE', 6),
    ('组织委员', 'NONE', 6),
    ('信息委员', 'NONE', 6),
]

def parse_position(text):
    if not text or text.strip() == '':
        return 'NONE', []
    
    matched = []
    for keyword, position, priority in POSITION_KEYWORDS:
        if keyword in text:
            if not any(m[0] == position for m in matched):
                matched.append((position, priority))
    
    if not matched:
        return 'NONE', []
    
    matched.sort(key=lambda x: x[1])
    return matched[0][0], [m[0] for m in matched[1:]]

async def main():
    db = Prisma()
    await db.connect()
    
    print('📖 读取大一名单...')
    wb1 = openpyxl.load_workbook(r'D:\myfile\clcode\大一名单.xlsx')
    ws1 = wb1.active
    
    students = {}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        sid = str(row[0]).strip()
        name = str(row[1]).strip()
        cls = str(row[2]).strip()
        students[(name, cls)] = sid
    
    print('📖 读取班委任职情况...')
    wb2 = openpyxl.load_workbook(r'D:\myfile\clcode\大一部分班级班级干部任命(2).xlsx')
    ws2 = wb2.active
    
    cadres = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        cls = str(row[1]).strip()
        name = str(row[2]).strip()
        position = str(row[5]).strip()
        cadres[(name, cls)] = position
    
    print('🔍 匹配班委信息...')
    matched_count = 0
    for (name, cls), position in cadres.items():
        if (name, cls) in students:
            matched_count += 1
    
    print('✅ 成功匹配 %d 名班委' % matched_count)
    
    # 创建密码
    hashed_password = bcrypt.hashpw('123456'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    # 导入学生
    created_count = 0
    skipped_count = 0
    cadre_count = 0
    
    print('\n📝 开始导入学生数据...')
    
    for (name, cls), sid in students.items():
        # 检查是否已存在
        existing = await db.user.find_unique(where={'studentId': sid})
        if existing:
            skipped_count += 1
            continue
        
        # 解析职位
        position_text = cadres.get((name, cls), '')
        primary_pos, secondary_pos = parse_position(position_text)
        
        # 创建用户
        user = await db.user.create(data={
            'name': name,
            'studentId': sid,
            'grade': 2025,
            'className': cls,
            'role': 'STUDENT',
            'phone': '',
            'password': hashed_password,
            'isFirstLogin': True,
        })
        
        # 创建报名记录
        await db.registration.create(data={
            'userId': user.id,
            'session': 'FIRST',
            'primaryPosition': primary_pos,
            'secondaryPositions': ','.join(secondary_pos) if secondary_pos else None,
            'status': 'APPROVED',
            'confirmedAt': None,
        })
        
        created_count += 1
        if primary_pos != 'NONE':
            cadre_count += 1
        
        if created_count % 50 == 0:
            print('  已导入 %d 人...' % created_count)
    
    print('\n✅ 导入完成！')
    print('  成功创建: %d 人' % created_count)
    print('  跳过(已存在): %d 人' % skipped_count)
    print('  其中班委: %d 人' % cadre_count)
    
    await db.disconnect()

if __name__ == '__main__':
    asyncio.run(main())
